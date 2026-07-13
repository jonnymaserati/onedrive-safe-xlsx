"""Round-trip tests for onedrive_safe_xlsx — no OneDrive/SharePoint needed.

We can't CI the SharePoint *outcome* (version lineage) — that needs a live tenant and is
validated manually. But we CAN test the *mechanism* the tool guarantees, which is the proxy
that outcome depends on:

  1. every non-data part openpyxl would drop (customXml, printerSettings, docProps/custom)
     survives in the output,
  2. docProps/core.xml has its change-tracking updated (dcterms:modified moved on,
     cp:revision bumped, cp:lastModifiedBy set),
  3. the openpyxl cell edit is present in the output.

The fixture is synthesised in-process: openpyxl writes a base workbook, then we inject the
non-data parts + a core.xml that carries revision/modified — i.e. a stand-in for a real
Excel/SharePoint-saved file that has the parts openpyxl strips on save.
"""
import io, os, re, tempfile, unittest, zipfile

import openpyxl

from onedrive_safe_xlsx import save_preserving

CORE_XML = (
    '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
    '<cp:coreProperties '
    'xmlns:cp="http://schemas.openxmlformats.org/package/2006/metadata/core-properties" '
    'xmlns:dcterms="http://purl.org/dc/terms/" '
    'xmlns:dc="http://purl.org/dc/elements/1.1/" '
    'xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">'
    '<dcterms:modified xsi:type="dcterms:W3CDTF">2020-01-01T00:00:00Z</dcterms:modified>'
    '<cp:lastModifiedBy>Original Author</cp:lastModifiedBy>'
    '<cp:revision>1</cp:revision>'
    '</cp:coreProperties>'
).encode("utf-8")

CUSTOM_XML = b'<?xml version="1.0"?><root><identity>keep-me</identity></root>'
PRINTER_BIN = b"\x00PRINTER-SETTINGS-BINARY\x00"

# non-data parts a real Excel file carries that openpyxl silently drops on save
EXTRA_PARTS = {
    "customXml/item1.xml": CUSTOM_XML,
    "xl/printerSettings/printerSettings1.bin": PRINTER_BIN,
}


def make_rich_xlsx(path):
    """A base openpyxl workbook, then injected with non-data parts + a revision-carrying
    core.xml — i.e. a file that HAS the parts openpyxl would strip."""
    wb = openpyxl.Workbook()
    wb.active["A1"] = "original"
    wb.save(path)
    with zipfile.ZipFile(path) as z:
        names = z.namelist()
        parts = {n: z.read(n) for n in names}
    parts["docProps/core.xml"] = CORE_XML   # overwrite with one that has revision/modified
    parts.update(EXTRA_PARTS)
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as z:
        for n, data in parts.items():
            z.writestr(n, data)


class SavePreservingTest(unittest.TestCase):
    def setUp(self):
        fd, self.path = tempfile.mkstemp(suffix=".xlsx"); os.close(fd)
        make_rich_xlsx(self.path)

    def tearDown(self):
        if os.path.exists(self.path):
            os.remove(self.path)

    def test_non_data_parts_preserved(self):
        wb = openpyxl.load_workbook(self.path)
        wb.active["A1"] = "edited"
        save_preserving(self.path, wb, modified_by="CI")
        with zipfile.ZipFile(self.path) as z:
            names = set(z.namelist())
            for part, expected in EXTRA_PARTS.items():
                self.assertIn(part, names, f"{part} was dropped")
                self.assertEqual(z.read(part), expected, f"{part} content changed")

    def test_core_change_tracking_updated(self):
        wb = openpyxl.load_workbook(self.path)
        wb.active["A1"] = "edited"
        save_preserving(self.path, wb, modified_by="CI User")
        with zipfile.ZipFile(self.path) as z:
            core = z.read("docProps/core.xml").decode("utf-8")
        self.assertNotIn("2020-01-01T00:00:00Z", core, "dcterms:modified was not moved on")
        self.assertIn("<cp:revision>2</cp:revision>", core, "cp:revision was not bumped")
        self.assertIn("<cp:lastModifiedBy>CI User</cp:lastModifiedBy>", core)

    def test_cell_edit_survives(self):
        wb = openpyxl.load_workbook(self.path)
        wb.active["A1"] = "edited"
        save_preserving(self.path, wb, modified_by="CI")
        wb2 = openpyxl.load_workbook(self.path)
        self.assertEqual(wb2.active["A1"].value, "edited")

    def test_out_param_leaves_original_untouched(self):
        fd, out = tempfile.mkstemp(suffix=".xlsx"); os.close(fd)
        try:
            wb = openpyxl.load_workbook(self.path)
            wb.active["A1"] = "edited"
            save_preserving(self.path, wb, out=out, modified_by="CI")
            self.assertEqual(openpyxl.load_workbook(self.path).active["A1"].value, "original")
            self.assertEqual(openpyxl.load_workbook(out).active["A1"].value, "edited")
        finally:
            if os.path.exists(out):
                os.remove(out)


if __name__ == "__main__":
    unittest.main()
